#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from statistics import mean, median, quantiles, stdev

import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
PALETTE = {
    "navy": "#16324F",
    "teal": "#2C7DA0",
    "gold": "#E0A458",
    "terracotta": "#D97D54",
    "sage": "#71917A",
    "ink": "#2D2A32",
    "paper": "#F7F3EB",
}
THREAD_COLORS = {
    1: PALETTE["navy"],
    4: PALETTE["teal"],
    8: PALETTE["gold"],
    16: PALETTE["terracotta"],
}
def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def parse_rows(path: Path) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for row in read_csv(path):
        rows.append(
            {
                "dim": int(row["dim"]),
                "N": int(row["N"]),
                "k": int(row["k"]),
                "mode": row["mode"],
                "threads": int(row["threads"]),
                "run_idx": int(row["run_idx"]),
                "iters": int(row["iters"]),
                "kernel_ms": float(row["kernel_ms"]),
                "total_ms": float(row["total_ms"]),
            }
        )
    return rows


def summarize(rows: list[dict[str, object]]) -> dict[str, list[dict[str, object]]]:
    grouped: dict[tuple[int, int, int, str, int], list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        key = (int(row["dim"]), int(row["N"]), int(row["k"]), str(row["mode"]), int(row["threads"]))
        grouped[key].append(row)

    mean_rows: list[dict[str, object]] = []
    for (dim, N, k, mode, threads), values in sorted(grouped.items()):
        kernel_values = [float(v["kernel_ms"]) for v in values]
        total_values = [float(v["total_ms"]) for v in values]
        iter_values = [int(v["iters"]) for v in values]
        mean_rows.append(
            {
                "dim": dim,
                "N": N,
                "k": k,
                "mode": mode,
                "threads": threads,
                "kernel_ms_mean": mean(kernel_values),
                "kernel_ms_std": stdev(kernel_values) if len(kernel_values) > 1 else 0.0,
                "total_ms_mean": mean(total_values),
                "total_ms_std": stdev(total_values) if len(total_values) > 1 else 0.0,
                "iters_mean": mean(iter_values),
                "iters_min": min(iter_values),
                "iters_max": max(iter_values),
            }
        )

    serial_lookup = {
        (int(row["dim"]), int(row["N"]), int(row["k"])): row
        for row in mean_rows
        if str(row["mode"]) == "serial"
    }

    speedup_rows: list[dict[str, object]] = []
    for row in mean_rows:
        if str(row["mode"]) != "omp":
            continue
        base = serial_lookup[(int(row["dim"]), int(row["N"]), int(row["k"]))]
        kernel_speedup = float(base["kernel_ms_mean"]) / float(row["kernel_ms_mean"])
        total_speedup = float(base["total_ms_mean"]) / float(row["total_ms_mean"])
        efficiency = kernel_speedup / int(row["threads"])
        speedup_rows.append(
            {
                "dim": int(row["dim"]),
                "N": int(row["N"]),
                "k": int(row["k"]),
                "threads": int(row["threads"]),
                "kernel_speedup": kernel_speedup,
                "total_speedup": total_speedup,
                "efficiency": efficiency,
                "kernel_ms_mean": float(row["kernel_ms_mean"]),
                "total_ms_mean": float(row["total_ms_mean"]),
                "serial_kernel_ms_mean": float(base["kernel_ms_mean"]),
                "serial_total_ms_mean": float(base["total_ms_mean"]),
            }
        )

    paired_rows: list[dict[str, object]] = []
    by_run: dict[tuple[int, int, int, int], dict[tuple[str, int], dict[str, object]]] = defaultdict(dict)
    for row in rows:
        by_run[(int(row["dim"]), int(row["N"]), int(row["k"]), int(row["run_idx"]))][
            (str(row["mode"]), int(row["threads"]))
        ] = row

    for (dim, N, k, run_idx), mapping in sorted(by_run.items()):
        serial_row = mapping.get(("serial", 1))
        if not serial_row:
            continue
        for threads in (1, 4, 8, 16):
            omp_row = mapping.get(("omp", threads))
            if not omp_row:
                continue
            paired_rows.append(
                {
                    "dim": dim,
                    "N": N,
                    "k": k,
                    "run_idx": run_idx,
                    "threads": threads,
                    "kernel_speedup_pair": float(serial_row["kernel_ms"]) / float(omp_row["kernel_ms"]),
                    "total_speedup_pair": float(serial_row["total_ms"]) / float(omp_row["total_ms"]),
                    "serial_kernel_ms": float(serial_row["kernel_ms"]),
                    "omp_kernel_ms": float(omp_row["kernel_ms"]),
                    "serial_total_ms": float(serial_row["total_ms"]),
                    "omp_total_ms": float(omp_row["total_ms"]),
                    "iters": int(serial_row["iters"]),
                }
            )

    robust_config_rows: list[dict[str, object]] = []
    pair_grouped: dict[tuple[int, int, int, int], list[dict[str, object]]] = defaultdict(list)
    for row in paired_rows:
        pair_grouped[(int(row["dim"]), int(row["N"]), int(row["k"]), int(row["threads"]))].append(row)

    for row in speedup_rows:
        key = (int(row["dim"]), int(row["N"]), int(row["k"]), int(row["threads"]))
        pair_values = [float(v["kernel_speedup_pair"]) for v in pair_grouped[key]]
        total_pair_values = [float(v["total_speedup_pair"]) for v in pair_grouped[key]]
        quartiles = quantiles(pair_values, n=4, method="inclusive")
        robust_config_rows.append(
            {
                "dim": int(row["dim"]),
                "N": int(row["N"]),
                "k": int(row["k"]),
                "threads": int(row["threads"]),
                "mean_based_kernel_speedup": float(row["kernel_speedup"]),
                "paired_mean_kernel_speedup": mean(pair_values),
                "paired_median_kernel_speedup": median(pair_values),
                "paired_q1_kernel_speedup": quartiles[0],
                "paired_q3_kernel_speedup": quartiles[2],
                "paired_iqr_kernel_speedup": quartiles[2] - quartiles[0],
                "paired_mean_total_speedup": mean(total_pair_values),
                "paired_median_total_speedup": median(total_pair_values),
                "mean_minus_median_gap": float(row["kernel_speedup"]) - median(pair_values),
            }
        )

    thread_summary: list[dict[str, object]] = []
    by_thread: dict[tuple[int, int], list[dict[str, object]]] = defaultdict(list)
    for row in speedup_rows:
        by_thread[(int(row["dim"]), int(row["threads"]))].append(row)
    by_thread_pairs: dict[tuple[int, int], list[dict[str, object]]] = defaultdict(list)
    for row in robust_config_rows:
        by_thread_pairs[(int(row["dim"]), int(row["threads"]))].append(row)
    for (dim, threads), values in sorted(by_thread.items()):
        robust_values = by_thread_pairs[(dim, threads)]
        thread_summary.append(
            {
                "dim": dim,
                "threads": threads,
                "avg_kernel_speedup": mean([float(v["kernel_speedup"]) for v in values]),
                "avg_total_speedup": mean([float(v["total_speedup"]) for v in values]),
                "avg_efficiency": mean([float(v["efficiency"]) for v in values]),
                "best_kernel_speedup": max(float(v["kernel_speedup"]) for v in values),
                "avg_paired_mean_speedup": mean([float(v["paired_mean_kernel_speedup"]) for v in robust_values]),
                "avg_paired_median_speedup": mean([float(v["paired_median_kernel_speedup"]) for v in robust_values]),
                "avg_paired_iqr": mean([float(v["paired_iqr_kernel_speedup"]) for v in robust_values]),
                "avg_mean_minus_median_gap": mean([float(v["mean_minus_median_gap"]) for v in robust_values]),
            }
        )

    best_thread_rows: list[dict[str, object]] = []
    grouped_by_dim_n: dict[tuple[int, int], list[dict[str, object]]] = defaultdict(list)
    for row in speedup_rows:
        grouped_by_dim_n[(int(row["dim"]), int(row["N"]))].append(row)
    for (dim, N), values in sorted(grouped_by_dim_n.items()):
        best = max(values, key=lambda item: float(item["kernel_speedup"]))
        best_thread_rows.append(
            {
                "dim": dim,
                "N": N,
                "best_threads": int(best["threads"]),
                "best_kernel_speedup": float(best["kernel_speedup"]),
                "best_total_speedup": float(best["total_speedup"]),
            }
        )

    robust_best_thread_rows: list[dict[str, object]] = []
    robust_grouped_by_dim_n: dict[tuple[int, int], list[dict[str, object]]] = defaultdict(list)
    for row in robust_config_rows:
        robust_grouped_by_dim_n[(int(row["dim"]), int(row["N"]))].append(row)
    for (dim, N), values in sorted(robust_grouped_by_dim_n.items()):
        best = max(values, key=lambda item: float(item["paired_median_kernel_speedup"]))
        robust_best_thread_rows.append(
            {
                "dim": dim,
                "N": N,
                "best_threads": int(best["threads"]),
                "paired_median_kernel_speedup": float(best["paired_median_kernel_speedup"]),
                "paired_q1_kernel_speedup": float(best["paired_q1_kernel_speedup"]),
                "paired_q3_kernel_speedup": float(best["paired_q3_kernel_speedup"]),
                "paired_iqr_kernel_speedup": float(best["paired_iqr_kernel_speedup"]),
            }
        )

    return {
        "mean_rows": mean_rows,
        "speedup_rows": speedup_rows,
        "paired_rows": paired_rows,
        "robust_config_rows": robust_config_rows,
        "thread_summary": thread_summary,
        "best_thread_rows": best_thread_rows,
        "robust_best_thread_rows": robust_best_thread_rows,
    }


def write_csv(path: Path, rows: list[dict[str, object]]) -> None:
    if not rows:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def style_axes(ax, title: str, xlabel: str, ylabel: str) -> None:
    ax.set_title(title, fontsize=15, fontweight="bold", color=PALETTE["ink"])
    ax.set_xlabel(xlabel, fontsize=11, color=PALETTE["ink"])
    ax.set_ylabel(ylabel, fontsize=11, color=PALETTE["ink"])
    ax.grid(True, alpha=0.22, color=PALETTE["navy"])
    ax.set_facecolor("white")
    for spine in ax.spines.values():
        spine.set_color("#D8D8D8")


def plot_speedup_overview(speedup_rows: list[dict[str, object]], out_path: Path) -> None:
    sns.set_theme(style="whitegrid", context="talk")
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.8), constrained_layout=True)
    for ax, dim in zip(axes, (2, 3)):
        rows = [row for row in speedup_rows if int(row["dim"]) == dim]
        by_n: dict[int, list[dict[str, object]]] = defaultdict(list)
        for row in rows:
            by_n[int(row["N"])].append(row)
        for N, values in sorted(by_n.items()):
            values = sorted(values, key=lambda item: int(item["threads"]))
            ax.plot(
                [int(v["threads"]) for v in values],
                [float(v["kernel_speedup"]) for v in values],
                marker="o",
                linewidth=2.4,
                markersize=6,
                label=f"N={N:,}".replace(",", ""),
            )
        style_axes(ax, f"Speedup en {dim}D", "Hilos", "Speedup (kernel)")
        ax.legend(fontsize=8, ncol=2, frameon=True)
    fig.patch.set_facecolor(PALETTE["paper"])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=220, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def plot_runtime_overview(mean_rows: list[dict[str, object]], out_path: Path) -> None:
    sns.set_theme(style="whitegrid", context="talk")
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.8), constrained_layout=True)
    for ax, dim in zip(axes, (2, 3)):
        rows = [row for row in mean_rows if int(row["dim"]) == dim]
        serial = sorted((row for row in rows if str(row["mode"]) == "serial"), key=lambda item: int(item["N"]))
        ax.plot(
            [int(r["N"]) for r in serial],
            [float(r["kernel_ms_mean"]) for r in serial],
            marker="o",
            linewidth=3,
            color=PALETTE["navy"],
            label="Serial",
        )
        for threads in (4, 8, 16):
            omp = sorted(
                (row for row in rows if str(row["mode"]) == "omp" and int(row["threads"]) == threads),
                key=lambda item: int(item["N"]),
            )
            ax.plot(
                [int(r["N"]) for r in omp],
                [float(r["kernel_ms_mean"]) for r in omp],
                marker="o",
                linewidth=2.3,
                color=THREAD_COLORS[threads],
                label=f"OpenMP {threads} hilos",
            )
        style_axes(ax, f"Tiempo promedio del kernel en {dim}D", "Número de puntos", "Tiempo promedio (ms)")
        ax.ticklabel_format(style="plain", axis="x")
        ax.legend(fontsize=8, frameon=True)
    fig.patch.set_facecolor(PALETTE["paper"])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=220, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def plot_runtime_median_overview(mean_rows: list[dict[str, object]], out_path: Path) -> None:
    sns.set_theme(style="whitegrid", context="talk")
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.8), constrained_layout=True)
    for ax, dim in zip(axes, (2, 3)):
        rows = [row for row in mean_rows if int(row["dim"]) == dim]
        serial = sorted((row for row in rows if str(row["mode"]) == "serial"), key=lambda item: int(item["N"]))
        ax.plot(
            [int(r["N"]) for r in serial],
            [float(r["kernel_ms_median"]) for r in serial],
            marker="o",
            linewidth=3,
            color=PALETTE["navy"],
            label="Serial",
        )
        omp_threads = sorted({int(row["threads"]) for row in rows if str(row["mode"]) == "omp"})
        for idx, threads in enumerate(omp_threads):
            omp = sorted(
                (row for row in rows if str(row["mode"]) == "omp" and int(row["threads"]) == threads),
                key=lambda item: int(item["N"]),
            )
            ax.plot(
                [int(r["N"]) for r in omp],
                [float(r["kernel_ms_median"]) for r in omp],
                marker="o",
                linewidth=2.3,
                color=thread_color(threads, idx + 1),
                label=f"OpenMP {threads} hilos",
            )
        style_axes(ax, f"Tiempo mediano del kernel en {dim}D", "Numero de puntos", "Tiempo mediano (ms)")
        ax.ticklabel_format(style="plain", axis="x")
        ax.legend(fontsize=8, frameon=True)
    fig.patch.set_facecolor(PALETTE["paper"])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=220, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def plot_kernel_cost_per_iter_overview(mean_rows: list[dict[str, object]], out_path: Path) -> None:
    sns.set_theme(style="whitegrid", context="talk")
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.8), constrained_layout=True)
    for ax, dim in zip(axes, (2, 3)):
        rows = [row for row in mean_rows if int(row["dim"]) == dim]
        serial = sorted((row for row in rows if str(row["mode"]) == "serial"), key=lambda item: int(item["N"]))
        ax.plot(
            [int(r["N"]) for r in serial],
            [float(r["kernel_ms_per_iter_mean"]) for r in serial],
            marker="o",
            linewidth=3,
            color=PALETTE["navy"],
            label="Serial",
        )
        omp_threads = sorted({int(row["threads"]) for row in rows if str(row["mode"]) == "omp"})
        for idx, threads in enumerate(omp_threads):
            omp = sorted(
                (row for row in rows if str(row["mode"]) == "omp" and int(row["threads"]) == threads),
                key=lambda item: int(item["N"]),
            )
            ax.plot(
                [int(r["N"]) for r in omp],
                [float(r["kernel_ms_per_iter_mean"]) for r in omp],
                marker="o",
                linewidth=2.3,
                color=thread_color(threads, idx + 1),
                label=f"OpenMP {threads} hilos",
            )
        style_axes(ax, f"Costo promedio por iteracion en {dim}D", "Numero de puntos", "Tiempo por iteracion (ms)")
        ax.ticklabel_format(style="plain", axis="x")
        ax.legend(fontsize=8, frameon=True)
    fig.patch.set_facecolor(PALETTE["paper"])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=220, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def plot_iteration_overview(mean_rows: list[dict[str, object]], out_path: Path) -> None:
    sns.set_theme(style="whitegrid", context="talk")
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.8), constrained_layout=True)
    for ax, dim in zip(axes, (2, 3)):
        serial = sorted(
            (row for row in mean_rows if int(row["dim"]) == dim and str(row["mode"]) == "serial"),
            key=lambda item: int(item["N"]),
        )
        xs = [int(row["N"]) for row in serial]
        medians = [float(row["iters_median"]) for row in serial]
        mins = [int(row["iters_min"]) for row in serial]
        maxs = [int(row["iters_max"]) for row in serial]
        ax.plot(xs, medians, marker="o", linewidth=3, color=PALETTE["navy"], label="Mediana de iteraciones")
        ax.fill_between(xs, mins, maxs, color=PALETTE["teal"], alpha=0.18, label="Rango min-max")
        style_axes(ax, f"Iteraciones por corrida en {dim}D", "Numero de puntos", "Iteraciones")
        ax.ticklabel_format(style="plain", axis="x")
        ax.legend(fontsize=8, frameon=True)
    fig.patch.set_facecolor(PALETTE["paper"])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=220, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def plot_efficiency_heatmaps(speedup_rows: list[dict[str, object]], out_path: Path) -> None:
    sns.set_theme(style="white", context="talk")
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.8), constrained_layout=True)
    for ax, dim in zip(axes, (2, 3)):
        rows = [row for row in speedup_rows if int(row["dim"]) == dim]
        Ns = sorted({int(row["N"]) for row in rows})
        threads = sorted({int(row["threads"]) for row in rows})
        matrix = []
        for N in Ns:
            line = []
            for thread in threads:
                match = next(row for row in rows if int(row["N"]) == N and int(row["threads"]) == thread)
                line.append(float(match["efficiency"]))
            matrix.append(line)
        sns.heatmap(
            matrix,
            ax=ax,
            cmap=sns.color_palette(["#F8F3EC", "#DCECEF", "#86BBD8", "#2C7DA0", "#16324F"], as_cmap=True),
            annot=True,
            fmt=".2f",
            cbar=True,
            xticklabels=threads,
            yticklabels=[str(N) for N in Ns],
        )
        ax.set_title(f"Eficiencia paralela en {dim}D", fontsize=15, fontweight="bold", color=PALETTE["ink"])
        ax.set_xlabel("Hilos")
        ax.set_ylabel("Número de puntos")
    fig.patch.set_facecolor(PALETTE["paper"])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=220, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def plot_convergence_overview(
    rows: list[dict[str, object]],
    thread_summary: list[dict[str, object]],
    out_path: Path,
) -> None:
    sns.set_theme(style="whitegrid", context="talk")
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.8), constrained_layout=True)

    total_rows = sorted(thread_summary, key=lambda item: (int(item["dim"]), int(item["threads"])))
    labels = [f"{row['dim']}D-{row['threads']}h" for row in total_rows]
    values = [float(row["avg_total_speedup"]) for row in total_rows]
    colors = [THREAD_COLORS[int(row["threads"])] for row in total_rows]
    axes[0].bar(labels, values, color=colors)
    style_axes(axes[0], "Speedup promedio usando total_ms", "Configuración", "Speedup promedio total")
    axes[0].tick_params(axis="x", rotation=40)

    serial_rows = [row for row in rows if str(row["mode"]) == "serial"]
    data = [
        [int(row["iters"]) for row in serial_rows if int(row["dim"]) == 2],
        [int(row["iters"]) for row in serial_rows if int(row["dim"]) == 3],
    ]
    bp = axes[1].boxplot(data, tick_labels=["2D", "3D"], patch_artist=True)
    for patch, color in zip(bp["boxes"], [PALETTE["teal"], PALETTE["gold"]]):
        patch.set_facecolor(color)
        patch.set_alpha(0.8)
    style_axes(axes[1], "Distribución de iteraciones hasta convergencia", "Dimensión", "Iteraciones")

    fig.patch.set_facecolor(PALETTE["paper"])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=220, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def plot_paired_speedup_boxplots(paired_rows: list[dict[str, object]], out_path: Path) -> None:
    sns.set_theme(style="whitegrid", context="talk")
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.8), constrained_layout=True)
    for ax, dim in zip(axes, (2, 3)):
        rows = [row for row in paired_rows if int(row["dim"]) == dim]
        data = []
        labels = []
        for threads in (1, 4, 8, 16):
            values = [float(row["kernel_speedup_pair"]) for row in rows if int(row["threads"]) == threads]
            data.append(values)
            labels.append(str(threads))
        bp = ax.boxplot(data, tick_labels=labels, patch_artist=True, showfliers=True)
        for patch, threads in zip(bp["boxes"], (1, 4, 8, 16)):
            patch.set_facecolor(THREAD_COLORS[threads])
            patch.set_alpha(0.75)
        style_axes(ax, f"Speedup emparejado por corrida en {dim}D", "Hilos", "Speedup por run_idx")
    fig.patch.set_facecolor(PALETTE["paper"])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=220, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def plot_mean_vs_median_comparison(thread_summary: list[dict[str, object]], out_path: Path) -> None:
    sns.set_theme(style="whitegrid", context="talk")
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.8), constrained_layout=True)
    for ax, dim in zip(axes, (2, 3)):
        rows = [row for row in thread_summary if int(row["dim"]) == dim]
        threads = [int(row["threads"]) for row in rows]
        mean_vals = [float(row["avg_kernel_speedup"]) for row in rows]
        median_vals = [float(row["avg_paired_median_speedup"]) for row in rows]
        x = range(len(threads))
        ax.bar([i - 0.18 for i in x], mean_vals, width=0.36, color=PALETTE["teal"], label="Promedio basado en medias")
        ax.bar([i + 0.18 for i in x], median_vals, width=0.36, color=PALETTE["gold"], label="Promedio de medianas emparejadas")
        ax.set_xticks(list(x))
        ax.set_xticklabels([str(t) for t in threads])
        style_axes(ax, f"Promedio vs mediana robusta en {dim}D", "Hilos", "Speedup resumido")
        ax.legend(fontsize=8, frameon=True)
    fig.patch.set_facecolor(PALETTE["paper"])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=220, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def plot_paired_median_trends(robust_config_rows: list[dict[str, object]], out_path: Path) -> None:
    sns.set_theme(style="whitegrid", context="talk")
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.8), constrained_layout=True)
    for ax, dim in zip(axes, (2, 3)):
        rows = [row for row in robust_config_rows if int(row["dim"]) == dim]
        for threads in (1, 4, 8, 16):
            values = sorted(
                (row for row in rows if int(row["threads"]) == threads),
                key=lambda item: int(item["N"]),
            )
            xs = [int(row["N"]) for row in values]
            medians = [float(row["paired_median_kernel_speedup"]) for row in values]
            q1 = [float(row["paired_q1_kernel_speedup"]) for row in values]
            q3 = [float(row["paired_q3_kernel_speedup"]) for row in values]
            ax.plot(
                xs,
                medians,
                marker="o",
                linewidth=2.4,
                markersize=6,
                color=THREAD_COLORS[threads],
                label=f"{threads} hilos",
            )
            ax.fill_between(xs, q1, q3, color=THREAD_COLORS[threads], alpha=0.16)
        style_axes(
            ax,
            f"Mediana emparejada e IQR en {dim}D",
            "Número de puntos",
            "Speedup emparejado del kernel",
        )
        ax.ticklabel_format(style="plain", axis="x")
        ax.legend(fontsize=8, frameon=True)
    fig.patch.set_facecolor(PALETTE["paper"])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=220, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def plot_stability_heatmaps(robust_config_rows: list[dict[str, object]], out_path: Path) -> None:
    sns.set_theme(style="white", context="talk")
    fig, axes = plt.subplots(1, 2, figsize=(14, 5.8), constrained_layout=True)
    for ax, dim in zip(axes, (2, 3)):
        rows = [row for row in robust_config_rows if int(row["dim"]) == dim]
        Ns = sorted({int(row["N"]) for row in rows})
        threads = sorted({int(row["threads"]) for row in rows})
        matrix = []
        for N in Ns:
            line = []
            for thread in threads:
                match = next(row for row in rows if int(row["N"]) == N and int(row["threads"]) == thread)
                line.append(float(match["paired_iqr_kernel_speedup"]))
            matrix.append(line)
        sns.heatmap(
            matrix,
            ax=ax,
            cmap=sns.color_palette(["#F8F3EC", "#F2D9C8", "#E0A458", "#D97D54", "#9B3D2D"], as_cmap=True),
            annot=True,
            fmt=".2f",
            cbar=True,
            xticklabels=threads,
            yticklabels=[str(N) for N in Ns],
        )
        ax.set_title(f"Variabilidad del speedup en {dim}D", fontsize=15, fontweight="bold", color=PALETTE["ink"])
        ax.set_xlabel("Hilos")
        ax.set_ylabel("Número de puntos")
    fig.patch.set_facecolor(PALETTE["paper"])
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out_path, dpi=220, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def create_workbook(
    output: Path,
    raw_rows: list[dict[str, object]],
    mean_rows: list[dict[str, object]],
    speedup_rows: list[dict[str, object]],
    paired_rows: list[dict[str, object]],
    robust_config_rows: list[dict[str, object]],
    thread_summary: list[dict[str, object]],
    best_thread_rows: list[dict[str, object]],
    robust_best_thread_rows: list[dict[str, object]],
    system_info: dict[str, str],
) -> None:
    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="D9E2F3")
    header_font = Font(bold=True, color="16324F")

    def write_sheet(name: str, rows: list[dict[str, object]]) -> None:
        ws = wb.create_sheet(title=name)
        headers = list(rows[0].keys()) if rows else []
        ws.append(headers)
        for row in rows:
            ws.append([row[h] for h in headers])
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for col in ws.columns:
            width = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = min(max(width, 12), 28)

    overview = wb.active
    overview.title = "Overview"
    overview["A1"] = "Workbook de análisis experimental de K-means paralelo"
    overview["A1"].font = Font(size=14, bold=True, color="16324F")
    overview["A3"] = "Hallazgo"
    overview["B3"] = "Valor"
    for cell in ("A3", "B3"):
        overview[cell].fill = header_fill
        overview[cell].font = header_font
    best_global = max(speedup_rows, key=lambda row: float(row["kernel_speedup"]))
    overview.append(["Mejor speedup global", round(float(best_global["kernel_speedup"]), 6)])
    overview.append(["Configuración", f"{best_global['dim']}D, N={best_global['N']}, {best_global['threads']} hilos"])
    best_robust = max(robust_config_rows, key=lambda row: float(row["paired_median_kernel_speedup"]))
    overview.append(["Mejor speedup robusto (mediana)", round(float(best_robust["paired_median_kernel_speedup"]), 6)])
    overview.append(["Configuración robusta", f"{best_robust['dim']}D, N={best_robust['N']}, {best_robust['threads']} hilos"])
    overview.append(["Filas experimentales", len(raw_rows)])
    overview.append(["CPU", system_info.get("cpu_name", "")])
    overview.append(["Sistema operativo", system_info.get("os_caption", "")])

    write_sheet("Raw_Experiments", raw_rows)
    write_sheet("Mean_By_Config", mean_rows)
    write_sheet("Speedup", speedup_rows)
    write_sheet("Paired_Speedup", paired_rows)
    write_sheet("Robust_Config", robust_config_rows)
    write_sheet("Thread_Summary", thread_summary)
    write_sheet("Best_Thread_By_N", best_thread_rows)
    write_sheet("Robust_Best_Thread_By_N", robust_best_thread_rows)

    ws_sys = wb.create_sheet(title="System_Info")
    ws_sys.append(["Clave", "Valor"])
    for key, value in system_info.items():
        ws_sys.append([key, value])
    for cell in ws_sys[1]:
        cell.fill = header_fill
        cell.font = header_font

    summary_ws = wb["Thread_Summary"]
    line_chart = LineChart()
    line_chart.title = "Speedup promedio del kernel por hilos"
    line_chart.y_axis.title = "Speedup promedio"
    line_chart.x_axis.title = "Filas del resumen"
    data = Reference(summary_ws, min_col=3, max_col=3, min_row=1, max_row=summary_ws.max_row)
    cats = Reference(summary_ws, min_col=2, max_col=2, min_row=2, max_row=summary_ws.max_row)
    line_chart.add_data(data, titles_from_data=True)
    data_med = Reference(summary_ws, min_col=7, max_col=7, min_row=1, max_row=summary_ws.max_row)
    line_chart.add_data(data_med, titles_from_data=True)
    line_chart.set_categories(cats)
    line_chart.height = 7
    line_chart.width = 12
    overview.add_chart(line_chart, "D3")

    bar_chart = BarChart()
    bar_chart.title = "Speedup promedio total por hilos"
    bar_chart.y_axis.title = "Speedup"
    data2 = Reference(summary_ws, min_col=4, max_col=4, min_row=1, max_row=summary_ws.max_row)
    bar_chart.add_data(data2, titles_from_data=True)
    bar_chart.set_categories(cats)
    bar_chart.height = 7
    bar_chart.width = 12
    overview.add_chart(bar_chart, "D20")

    robust_ws = wb["Robust_Best_Thread_By_N"]
    robust_chart = LineChart()
    robust_chart.title = "Mejor speedup robusto por tamaño"
    robust_chart.y_axis.title = "Mediana emparejada"
    robust_chart.x_axis.title = "Filas del resumen robusto"
    robust_data = Reference(robust_ws, min_col=3, max_col=3, min_row=1, max_row=robust_ws.max_row)
    robust_cats = Reference(robust_ws, min_col=2, max_col=2, min_row=2, max_row=robust_ws.max_row)
    robust_chart.add_data(robust_data, titles_from_data=True)
    robust_chart.set_categories(robust_cats)
    robust_chart.height = 7
    robust_chart.width = 12
    overview.add_chart(robust_chart, "D37")

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate professional analysis assets for the K-means report.")
    parser.add_argument("--experiments", default=str(ROOT / "results" / "experiments.csv"))
    parser.add_argument("--system-info", default=str(ROOT / "results" / "system_info.txt"))
    parser.add_argument("--outdir", default=str(ROOT / "results"))
    args = parser.parse_args()

    outdir = Path(args.outdir)
    rows = parse_rows(Path(args.experiments))
    system_info = {}
    for line in Path(args.system_info).read_text(encoding="utf-8").splitlines():
        if ":" in line:
            key, value = line.split(":", 1)
            system_info[key.strip()] = value.strip()

    summary = summarize(rows)
    write_csv(outdir / "analysis_mean_by_config.csv", summary["mean_rows"])
    write_csv(outdir / "analysis_speedup_by_config.csv", summary["speedup_rows"])
    write_csv(outdir / "analysis_pairwise_speedup.csv", summary["paired_rows"])
    write_csv(outdir / "analysis_robust_config_summary.csv", summary["robust_config_rows"])
    write_csv(outdir / "analysis_thread_summary.csv", summary["thread_summary"])
    write_csv(outdir / "analysis_best_thread_by_n.csv", summary["best_thread_rows"])
    write_csv(outdir / "analysis_robust_best_thread_by_n.csv", summary["robust_best_thread_rows"])

    plot_speedup_overview(summary["speedup_rows"], outdir / "report_speedup_overview.png")
    plot_runtime_overview(summary["mean_rows"], outdir / "report_kernel_runtime_overview.png")
    plot_efficiency_heatmaps(summary["speedup_rows"], outdir / "report_efficiency_heatmaps.png")
    plot_convergence_overview(rows, summary["thread_summary"], outdir / "report_convergence_overview.png")
    plot_paired_speedup_boxplots(summary["paired_rows"], outdir / "report_paired_speedup_boxplots.png")
    plot_mean_vs_median_comparison(summary["thread_summary"], outdir / "report_mean_vs_median_comparison.png")
    plot_paired_median_trends(summary["robust_config_rows"], outdir / "report_paired_median_trends.png")
    plot_stability_heatmaps(summary["robust_config_rows"], outdir / "report_stability_heatmaps.png")

    create_workbook(
        outdir / "analysis_workbook.xlsx",
        rows,
        summary["mean_rows"],
        summary["speedup_rows"],
        summary["paired_rows"],
        summary["robust_config_rows"],
        summary["thread_summary"],
        summary["best_thread_rows"],
        summary["robust_best_thread_rows"],
        system_info,
    )

    print(f"Wrote {outdir / 'analysis_mean_by_config.csv'}")
    print(f"Wrote {outdir / 'analysis_speedup_by_config.csv'}")
    print(f"Wrote {outdir / 'analysis_pairwise_speedup.csv'}")
    print(f"Wrote {outdir / 'analysis_robust_config_summary.csv'}")
    print(f"Wrote {outdir / 'analysis_thread_summary.csv'}")
    print(f"Wrote {outdir / 'analysis_best_thread_by_n.csv'}")
    print(f"Wrote {outdir / 'analysis_robust_best_thread_by_n.csv'}")
    print(f"Wrote {outdir / 'report_speedup_overview.png'}")
    print(f"Wrote {outdir / 'report_kernel_runtime_overview.png'}")
    print(f"Wrote {outdir / 'report_efficiency_heatmaps.png'}")
    print(f"Wrote {outdir / 'report_convergence_overview.png'}")
    print(f"Wrote {outdir / 'report_paired_speedup_boxplots.png'}")
    print(f"Wrote {outdir / 'report_mean_vs_median_comparison.png'}")
    print(f"Wrote {outdir / 'report_paired_median_trends.png'}")
    print(f"Wrote {outdir / 'report_stability_heatmaps.png'}")
    print(f"Wrote {outdir / 'analysis_workbook.xlsx'}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
